"""Proxy Engine Exporters - Multi-Format Export System"""

import csv
import json
import logging
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
from urllib.parse import urlparse

# Third-party imports with safe fallbacks
try:
    import yaml
    HAS_YAML = True
except ImportError:
    HAS_YAML = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Import our models (using exact names from previous files)
from ..proxy_core.models import ProxyInfo, ProxyProtocol, AnonymityLevel, ThreatLevel, ProxyStatus
from ..proxy_core.config import ConfigManager
from .analysis.geo_analyzer import GeoAnalyzer


# ===============================================================================
# EXPORT CONFIGURATION CLASSES
# ===============================================================================

class ExportFormat(Enum):
    """Supported export formats"""
    JSON = "json"
    CSV = "csv"
    TXT = "txt"
    XLSX = "xlsx"
    YAML = "yaml"


@dataclass
class ExportConfig:
    """Configuration for export operations"""
    include_headers: bool = True
    include_metadata: bool = True
    include_metrics: bool = True
    include_geo_data: bool = True
    include_security_data: bool = True
    pretty_format: bool = True
    timestamp_format: str = "%Y-%m-%d %H:%M:%S"
    
    # Field selection
    basic_fields: List[str] = field(default_factory=lambda: [
        "proxy_id", "host", "port", "protocol", "status"
    ])
    extended_fields: List[str] = field(default_factory=lambda: [
        "anonymity_level", "threat_level", "security_level", "source", "discovered_at"
    ])
    metrics_fields: List[str] = field(default_factory=lambda: [
        "response_time", "success_rate", "uptime_percentage", "check_count"
    ])
    geo_fields: List[str] = field(default_factory=lambda: [
        "country", "country_code", "city", "region", "isp", "asn"
    ])


@dataclass
class ExportResult:
    """Result of export operation"""
    format_type: ExportFormat
    file_path: str
    proxy_count: int
    file_size_bytes: int
    export_duration: float
    success: bool
    error_message: Optional[str] = None
    timestamp: datetime = field(default_factory=datetime.utcnow)
    
    @property
    def file_size_mb(self) -> float:
        return self.file_size_bytes / (1024 * 1024)


# ===============================================================================
# BASE EXPORTER CLASS
# ===============================================================================

class BaseExporter(ABC):
    """Abstract base class for all exporters"""
    
    def __init__(self, config: Optional[ConfigManager] = None, 
                 export_config: Optional[ExportConfig] = None):
        self.config = config or ConfigManager()
        self.export_config = export_config or ExportConfig()
        self.logger = logging.getLogger(__name__)
        
        # Initialize geo analyzer for fallback geographic detection
        self.geo_analyzer = GeoAnalyzer(config) if self.export_config.include_geo_data else None
    
    @abstractmethod
    def export(self, proxies: List[ProxyInfo], file_path: str) -> ExportResult:
        """Export proxies to file"""
        pass
    
    def _prepare_proxy_data(self, proxies: List[ProxyInfo]) -> List[Dict[str, Any]]:
        """Prepare proxy data for export with optional threading"""
        if len(proxies) > 200:
            # Use threaded preparation for large datasets
            return self._prepare_proxy_data_threaded(proxies)
        else:
            # Use sequential preparation for smaller datasets
            return self._prepare_proxy_data_sequential(proxies)
    
    def _prepare_proxy_data_sequential(self, proxies: List[ProxyInfo]) -> List[Dict[str, Any]]:
        """Sequential proxy data preparation (original implementation)"""
        prepared_data = []
        
        for proxy in proxies:
            data = {}
            
            # Basic fields
            for field in self.export_config.basic_fields:
                if hasattr(proxy, field):
                    value = getattr(proxy, field)
                    if hasattr(value, 'value'):  # Handle Enums
                        data[field] = value.value
                    elif isinstance(value, datetime):
                        data[field] = value.strftime(self.export_config.timestamp_format)
                    else:
                        data[field] = value
            
            # Extended fields
            for field in self.export_config.extended_fields:
                if hasattr(proxy, field):
                    value = getattr(proxy, field)
                    if hasattr(value, 'value'):  # Handle Enums
                        data[field] = value.value
                    elif isinstance(value, datetime):
                        data[field] = value.strftime(self.export_config.timestamp_format)
                    else:
                        data[field] = value
            
            # Add proxy URL and address
            data['address'] = proxy.address
            data['url'] = proxy.url
            data['is_authenticated'] = proxy.is_authenticated
            
            # Metrics data
            if self.export_config.include_metrics and proxy.metrics:
                for field in self.export_config.metrics_fields:
                    if hasattr(proxy.metrics, field):
                        value = getattr(proxy.metrics, field)
                        if isinstance(value, datetime) and value:
                            data[f"metrics_{field}"] = value.strftime(self.export_config.timestamp_format)
                        else:
                            data[f"metrics_{field}"] = value
            
            # Geographic data
            if self.export_config.include_geo_data and proxy.geo_location:
                for field in self.export_config.geo_fields:
                    if hasattr(proxy.geo_location, field):
                        value = getattr(proxy.geo_location, field)
                        data[f"geo_{field}"] = value
            
            # Tags and notes
            data['tags'] = ', '.join(proxy.tags) if proxy.tags else ''
            data['notes'] = proxy.notes
            
            prepared_data.append(data)
        
        return prepared_data
    
    def _prepare_proxy_data_threaded(self, proxies: List[ProxyInfo]) -> List[Dict[str, Any]]:
        """Optimized threaded proxy data preparation for large datasets"""
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading
        import os
        
        # Optimize chunk size based on dataset size and CPU cores
        cpu_count = os.cpu_count() or 4
        optimal_chunk_size = max(25, len(proxies) // (cpu_count * 2))
        chunks = [proxies[i:i + optimal_chunk_size] for i in range(0, len(proxies), optimal_chunk_size)]
        
        def process_chunk(chunk):
            """Process a chunk of proxies with error handling"""
            try:
                return self._prepare_proxy_data_sequential(chunk)
            except Exception as e:
                self.logger.warning(f"Chunk processing failed: {e}")
                return []
        
        # Use optimal number of workers
        max_workers = min(cpu_count, len(chunks), 8)  # Cap at 8 for memory efficiency
        prepared_data = []
        
        try:
            with ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="ProxyExporter") as executor:
                # Submit all chunks
                future_to_chunk = {executor.submit(process_chunk, chunk): i for i, chunk in enumerate(chunks)}
                
                # Collect results in order for consistent output
                chunk_results = [None] * len(chunks)
                completed_chunks = 0
                
                for future in as_completed(future_to_chunk):
                    chunk_index = future_to_chunk[future]
                    try:
                        chunk_result = future.result(timeout=30)  # Add timeout
                        chunk_results[chunk_index] = chunk_result or []
                        completed_chunks += 1
                        
                        # Log progress for large datasets
                        if len(proxies) > 1000 and completed_chunks % 5 == 0:
                            self.logger.debug(f"Processed {completed_chunks}/{len(chunks)} chunks")
                            
                    except Exception as e:
                        self.logger.error(f"Chunk {chunk_index} failed: {e}")
                        chunk_results[chunk_index] = []
                
                # Combine all results efficiently
                for chunk_result in chunk_results:
                    if chunk_result:
                        prepared_data.extend(chunk_result)
                        
        except Exception as e:
            self.logger.error(f"Threaded preparation failed: {e}")
            # Fallback to sequential processing
            return self._prepare_proxy_data_sequential(proxies)
        
        return prepared_data
    
    def _get_file_size(self, file_path: str) -> int:
        """Get file size in bytes"""
        try:
            return Path(file_path).stat().st_size
        except Exception:
            return 0
    
    def _create_metadata(self, proxies: List[ProxyInfo]) -> Dict[str, Any]:
        """Create export metadata"""
        return {
            'export_timestamp': datetime.utcnow().isoformat(),
            'proxy_count': len(proxies),
            'export_config': {
                'include_metadata': self.export_config.include_metadata,
                'include_metrics': self.export_config.include_metrics,
                'include_geo_data': self.export_config.include_geo_data,
                'include_security_data': self.export_config.include_security_data
            },
            'summary': self._generate_summary(proxies)
        }
    
    def _generate_summary(self, proxies: List[ProxyInfo]) -> Dict[str, Any]:
        """Generate summary statistics"""
        if not proxies:
            return {}
        
        # Count by protocol
        protocol_counts = {}
        for protocol in ProxyProtocol:
            protocol_counts[protocol.value] = sum(1 for p in proxies if p.protocol == protocol)
        
        # Count by anonymity level
        anonymity_counts = {}
        for level in AnonymityLevel:
            anonymity_counts[level.value] = sum(1 for p in proxies if p.anonymity_level == level)
        
        # Count by status
        status_counts = {}
        for status in ProxyStatus:
            status_counts[status.value] = sum(1 for p in proxies if p.status == status)
        
        # Performance statistics
        response_times = [p.metrics.response_time for p in proxies 
                         if p.metrics.response_time is not None]
        success_rates = [p.metrics.success_rate for p in proxies 
                        if p.metrics.success_rate is not None]
        
        return {
            'total_proxies': len(proxies),
            'protocol_distribution': protocol_counts,
            'anonymity_distribution': anonymity_counts,
            'status_distribution': status_counts,
            'performance_stats': {
                'avg_response_time': sum(response_times) / len(response_times) if response_times else None,
                'avg_success_rate': sum(success_rates) / len(success_rates) if success_rates else None,
                'fastest_proxy': min(response_times) if response_times else None,
                'slowest_proxy': max(response_times) if response_times else None
            }
        }
    
    def _ensure_geo_data(self, proxy: ProxyInfo) -> None:
        """Ensure proxy has geographic data, perform analysis if missing"""
        if not self.geo_analyzer:
            return
        
        if not proxy.geo_location or not proxy.geo_location.country_code:
            try:
                geo_result = self.geo_analyzer.analyze_proxy(proxy)
                if geo_result.success:
                    self.logger.debug(f"Fallback geographic analysis completed for {proxy.host}")
            except Exception as e:
                self.logger.debug(f"Fallback geographic analysis failed for {proxy.host}: {e}")


# ===============================================================================
# JSON EXPORTER
# ===============================================================================

class JSONExporter(BaseExporter):
    """JSON format exporter"""
    
    def export(self, proxies: List[ProxyInfo], file_path: str) -> ExportResult:
        """Export proxies to JSON file"""
        import time
        start_time = time.time()
        
        result = ExportResult(
            format_type=ExportFormat.JSON,
            file_path=file_path,
            proxy_count=len(proxies),
            file_size_bytes=0,
            export_duration=0.0,
            success=False
        )
        
        try:
            # Prepare data
            proxy_data = self._prepare_proxy_data(proxies)
            
            # Create export structure
            export_data = {
                'proxies': proxy_data
            }
            
            # Add metadata if requested
            if self.export_config.include_metadata:
                export_data['metadata'] = self._create_metadata(proxies)
            
            # Write to file
            with open(file_path, 'w', encoding='utf-8') as f:
                if self.export_config.pretty_format:
                    json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
                else:
                    json.dump(export_data, f, ensure_ascii=False, default=str)
            
            result.file_size_bytes = self._get_file_size(file_path)
            result.success = True
            
        except Exception as e:
            result.error_message = str(e)
            self.logger.error(f"JSON export failed: {e}")
        
        finally:
            result.export_duration = time.time() - start_time
        
        return result


# ===============================================================================
# CSV EXPORTER
# ===============================================================================

class CSVExporter(BaseExporter):
    """CSV format exporter"""
    
    def export(self, proxies: List[ProxyInfo], file_path: str) -> ExportResult:
        """Export proxies to CSV file"""
        import time
        start_time = time.time()
        
        result = ExportResult(
            format_type=ExportFormat.CSV,
            file_path=file_path,
            proxy_count=len(proxies),
            file_size_bytes=0,
            export_duration=0.0,
            success=False
        )
        
        try:
            # Prepare data
            proxy_data = self._prepare_proxy_data(proxies)
            
            if not proxy_data:
                result.error_message = "No proxy data to export"
                return result
            
            # Get all possible field names
            all_fields = set()
            for data in proxy_data:
                all_fields.update(data.keys())
            
            # Sort fields for consistent output
            sorted_fields = sorted(all_fields)
            
            # Write CSV file
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=sorted_fields)
                
                if self.export_config.include_headers:
                    writer.writeheader()
                
                writer.writerows(proxy_data)
            
            result.file_size_bytes = self._get_file_size(file_path)
            result.success = True
            
        except Exception as e:
            result.error_message = str(e)
            self.logger.error(f"CSV export failed: {e}")
        
        finally:
            result.export_duration = time.time() - start_time
        
        return result


# ===============================================================================
# TXT EXPORTER
# ===============================================================================

class TXTExporter(BaseExporter):
    """Plain text format exporter"""
    
    def export(self, proxies: List[ProxyInfo], file_path: str) -> ExportResult:
        """Export proxies to text file"""
        import time
        start_time = time.time()
        
        result = ExportResult(
            format_type=ExportFormat.TXT,
            file_path=file_path,
            proxy_count=len(proxies),
            file_size_bytes=0,
            export_duration=0.0,
            success=False
        )
        
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                # Write header if requested
                if self.export_config.include_headers:
                    f.write("# Proxy Hunter Export\n")
                    f.write(f"# Generated: {datetime.utcnow().strftime(self.export_config.timestamp_format)}\n")
                    f.write(f"# Total Proxies: {len(proxies)}\n")
                    f.write("# Format: host:port or protocol://host:port\n\n")
                
                # Write proxies with improved error handling
                for proxy in proxies:
                    try:
                        if self.export_config.include_metadata:
                            # Ensure geo data is available for country information
                            self._ensure_geo_data(proxy)
                            
                            # Extended format with metadata
                            line = f"{proxy.url}"
                            
                            # Add response time if available
                            if proxy.metrics and proxy.metrics.response_time:
                                line += f" {proxy.metrics.response_time:.0f}ms"
                            
                            # Add anonymity level
                            if proxy.anonymity_level:
                                line += f" {proxy.anonymity_level.value}"
                            
                            # Add country code if available
                            if proxy.geo_location and proxy.geo_location.country_code:
                                line += f" {proxy.geo_location.country_code}"
                            
                            f.write(line + "\n")
                        else:
                            # Simple format with validation
                            address = getattr(proxy, 'address', f"{proxy.host}:{proxy.port}")
                            f.write(f"{address}\n")
                    except Exception as e:
                        self.logger.warning(f"Failed to write proxy {proxy.proxy_id}: {e}")
                        # Write minimal fallback format
                        f.write(f"{proxy.host}:{proxy.port}\n")
            
            result.file_size_bytes = self._get_file_size(file_path)
            result.success = True
            
        except Exception as e:
            result.error_message = str(e)
            self.logger.error(f"TXT export failed: {e}")
        
        finally:
            result.export_duration = time.time() - start_time
        
        return result


# ===============================================================================
# XLSX EXPORTER
# ===============================================================================

class XLSXExporter(BaseExporter):
    """Excel XLSX format exporter"""
    
    def export(self, proxies: List[ProxyInfo], file_path: str) -> ExportResult:
        """Export proxies to Excel file"""
        if not HAS_OPENPYXL:
            return ExportResult(
                format_type=ExportFormat.XLSX,
                file_path=file_path,
                proxy_count=0,
                file_size_bytes=0,
                export_duration=0.0,
                success=False,
                error_message="openpyxl library required for XLSX export"
            )
        
        import time
        start_time = time.time()
        
        result = ExportResult(
            format_type=ExportFormat.XLSX,
            file_path=file_path,
            proxy_count=len(proxies),
            file_size_bytes=0,
            export_duration=0.0,
            success=False
        )
        
        try:
            # Create workbook
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Proxies"
            
            # Prepare data
            proxy_data = self._prepare_proxy_data(proxies)
            
            if not proxy_data:
                result.error_message = "No proxy data to export"
                return result
            
            # Get headers
            headers = list(proxy_data[0].keys())
            
            # Write headers with formatting
            if self.export_config.include_headers:
                for col, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            # Write data
            for row, proxy_dict in enumerate(proxy_data, 2):
                for col, header in enumerate(headers, 1):
                    value = proxy_dict.get(header, '')
                    worksheet.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary sheet if metadata included
            if self.export_config.include_metadata:
                self._add_summary_sheet(workbook, proxies)
            
            # Save workbook
            workbook.save(file_path)
            
            result.file_size_bytes = self._get_file_size(file_path)
            result.success = True
            
        except Exception as e:
            result.error_message = str(e)
            self.logger.error(f"XLSX export failed: {e}")
        
        finally:
            result.export_duration = time.time() - start_time
        
        return result
    
    def _add_summary_sheet(self, workbook, proxies: List[ProxyInfo]):
        """Add summary statistics sheet"""
        summary_sheet = workbook.create_sheet("Summary")
        
        # Get summary data
        summary = self._generate_summary(proxies)
        
        row = 1
        
        # Basic statistics
        summary_sheet.cell(row=row, column=1, value="Export Summary").font = Font(bold=True, size=14)
        row += 2
        
        summary_sheet.cell(row=row, column=1, value="Total Proxies:")
        summary_sheet.cell(row=row, column=2, value=summary['total_proxies'])
        row += 1
        
        summary_sheet.cell(row=row, column=1, value="Export Time:")
        summary_sheet.cell(row=row, column=2, value=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"))
        row += 2
        
        # Protocol distribution
        summary_sheet.cell(row=row, column=1, value="Protocol Distribution").font = Font(bold=True)
        row += 1
        
        for protocol, count in summary['protocol_distribution'].items():
            summary_sheet.cell(row=row, column=1, value=protocol)
            summary_sheet.cell(row=row, column=2, value=count)
            row += 1
        
        row += 1
        
        # Anonymity distribution
        summary_sheet.cell(row=row, column=1, value="Anonymity Distribution").font = Font(bold=True)
        row += 1
        
        for level, count in summary['anonymity_distribution'].items():
            summary_sheet.cell(row=row, column=1, value=level)
            summary_sheet.cell(row=row, column=2, value=count)
            row += 1


# ===============================================================================
# YAML EXPORTER
# ===============================================================================

class YAMLExporter(BaseExporter):
    """YAML format exporter"""
    
    def export(self, proxies: List[ProxyInfo], file_path: str) -> ExportResult:
        """Export proxies to YAML file"""
        if not HAS_YAML:
            return ExportResult(
                format_type=ExportFormat.YAML,
                file_path=file_path,
                proxy_count=0,
                file_size_bytes=0,
                export_duration=0.0,
                success=False,
                error_message="PyYAML library required for YAML export"
            )
        
        import time
        start_time = time.time()
        
        result = ExportResult(
            format_type=ExportFormat.YAML,
            file_path=file_path,
            proxy_count=len(proxies),
            file_size_bytes=0,
            export_duration=0.0,
            success=False
        )
        
        try:
            # Prepare data
            proxy_data = self._prepare_proxy_data(proxies)
            
            # Create export structure
            export_data = {
                'proxies': proxy_data
            }
            
            # Add metadata if requested
            if self.export_config.include_metadata:
                export_data['metadata'] = self._create_metadata(proxies)
            
            # Write to file
            with open(file_path, 'w', encoding='utf-8') as f:
                yaml.dump(export_data, f, default_flow_style=False, 
                         allow_unicode=True, indent=2)
            
            result.file_size_bytes = self._get_file_size(file_path)
            result.success = True
            
        except Exception as e:
            result.error_message = str(e)
            self.logger.error(f"YAML export failed: {e}")
        
        finally:
            result.export_duration = time.time() - start_time
        
        return result


# ===============================================================================
# MAIN EXPORT MANAGER
# ===============================================================================

class ExportManager:
    """Main export manager orchestrating all export operations"""
    
    def __init__(self, config: Optional[ConfigManager] = None, 
                 export_config: Optional[ExportConfig] = None):
        self.config = config or ConfigManager()
        self.export_config = export_config or ExportConfig()
        self.logger = logging.getLogger(__name__)
        
        # Initialize exporters
        self.exporters = {
            ExportFormat.JSON: JSONExporter(config, export_config),
            ExportFormat.CSV: CSVExporter(config, export_config),
            ExportFormat.TXT: TXTExporter(config, export_config),
            ExportFormat.XLSX: XLSXExporter(config, export_config),
            ExportFormat.YAML: YAMLExporter(config, export_config)
        }
    
    def export_proxies(self, proxies: List[ProxyInfo], file_path: str, 
                      format_type: Optional[Union[ExportFormat, str]] = None) -> ExportResult:
        """Export proxies to specified format"""
        
        # Determine format from file extension if not specified
        if format_type is None:
            format_type = self._detect_format_from_path(file_path)
        
        # Convert string to enum if needed
        if isinstance(format_type, str):
            try:
                format_type = ExportFormat(format_type.lower())
            except ValueError:
                return ExportResult(
                    format_type=ExportFormat.JSON,
                    file_path=file_path,
                    proxy_count=0,
                    file_size_bytes=0,
                    export_duration=0.0,
                    success=False,
                    error_message=f"Unsupported format: {format_type}"
                )
        
        # Get appropriate exporter
        exporter = self.exporters.get(format_type)
        if not exporter:
            return ExportResult(
                format_type=format_type,
                file_path=file_path,
                proxy_count=0,
                file_size_bytes=0,
                export_duration=0.0,
                success=False,
                error_message=f"No exporter available for {format_type.value}"
            )
        
        # Perform export
        self.logger.info(f"Exporting {len(proxies)} proxies to {format_type.value} format")
        result = exporter.export(proxies, file_path)
        
        if result.success:
            self.logger.info(f"Export completed: {file_path} ({result.file_size_mb:.2f} MB)")
        else:
            self.logger.error(f"Export failed: {result.error_message}")
        
        return result
    
    def _detect_format_from_path(self, file_path: str) -> ExportFormat:
        """Detect export format from file extension"""
        path = Path(file_path)
        extension = path.suffix.lower()
        
        format_map = {
            '.json': ExportFormat.JSON,
            '.csv': ExportFormat.CSV,
            '.txt': ExportFormat.TXT,
            '.xlsx': ExportFormat.XLSX,
            '.xls': ExportFormat.XLSX,
            '.yaml': ExportFormat.YAML,
            '.yml': ExportFormat.YAML
        }
        
        return format_map.get(extension, ExportFormat.JSON)
    
    def export_multiple_formats(self, proxies: List[ProxyInfo], 
                               base_path: str, 
                               formats: List[ExportFormat]) -> Dict[ExportFormat, ExportResult]:
        """Export to multiple formats"""
        results = {}
        base_name = Path(base_path).stem
        base_dir = Path(base_path).parent
        
        for format_type in formats:
            file_path = base_dir / f"{base_name}.{format_type.value}"
            result = self.export_proxies(proxies, str(file_path), format_type)
            results[format_type] = result
        
        return results
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported export formats"""
        return [fmt.value for fmt in ExportFormat]


# ===============================================================================
# UTILITY FUNCTIONS
# ===============================================================================

def create_export_manager(config: Optional[ConfigManager] = None,
                         include_metadata: bool = True,
                         include_metrics: bool = True) -> ExportManager:
    """Create export manager with specified configuration"""
    export_config = ExportConfig(
        include_metadata=include_metadata,
        include_metrics=include_metrics
    )
    return ExportManager(config, export_config)


def create_json_exporter(config: Optional[ConfigManager] = None) -> JSONExporter:
    """Create JSON exporter"""
    return JSONExporter(config)


def create_csv_exporter(config: Optional[ConfigManager] = None) -> CSVExporter:
    """Create CSV exporter"""
    return CSVExporter(config)


def export_proxy_list(proxies: List[ProxyInfo], file_path: str, 
                     format_type: Optional[str] = None) -> bool:
    """Quick utility function to export proxy list"""
    try:
        manager = create_export_manager()
        result = manager.export_proxies(proxies, file_path, format_type)
        return result.success
    except Exception:
        return False


# ===============================================================================
# PROXY EXPORTER ALIAS
# ===============================================================================

# Alias for compatibility with __init__.py
ProxyExporter = ExportManager